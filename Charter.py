import os
import sys
import orionsdk
import getpass
import socket
import netmiko
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pyfiglet import Figlet
class Query():
	def Select(DB):
		Results = DB.query("SELECT Caption,IPAddress FROM Orion.Nodes WHERE IPAddress LIKE '10.255.254%'")
		CIP = {}
		for Keys in Results['results']:
			C = Keys['Caption']
			IP = Keys['IPAddress']
			CIP[C] = IP
		return CIP
	def FindOccur(file,string):
		fhand = open(file,"r+")
		count = 0
		for line in fhand:
			line = line.strip()
			print(line)
			print("--- Searching for string,",string)
			if string in line:
				print("Match Found")
				print(count)
				count = count + 1
			else:
				continue
		print("Returning Count",count)
		return count
	def TempFileMkr(Fname,data):
		fhand = open(Fname + ".txt","w+")
		fhand.write(data)
		fhand.close()
		fname = Fname + ".txt"
		return fname
	def Binder(MN,IP,AP,NP,DP,TP):
		ID = {
		"MACHINE_NAME" : MN,
		"IP_ADDRESS" : IP,
		"ACTIVE_PORTS" : AP,
		"NON_ACTIVE_PORTS" : NP,
		"DISABLED_PORTS" : DP,
		"TOTAL_PORTS" : TP
		}
		return ID
# Connection class facilitates the SSH tunnel that enables commands
class Connection():
	def TestConnectAbility(IP,username,password):
		Test = {
		"device_type":"cisco_ios",
		"ip":IP,
		"username":username,
		"password":password
		}
		# If unable to connect, the connection will return a False which data values will = N/A
		try:
			connection = netmiko.ConnectHandler(**Test)
			return True,connection
		except:
			print("No connection")
			connection = None
			return False,connection
	# Returns Data
	def Pull(Connection,cmd):
		output = Connection.send_command(cmd)
		return output




# Device is a facilitator for SSH CLI connections, this is where commands return relevant data

			


class Main():
	def __init__(self):
		self.username = getpass.getpass()
		self.password = getpass.getpass()
		self.DB = orionsdk.SwisClient("172.16.3.57",self.username,self.password)
		self.Command = "show interfaces status"
		self.SettupXL()
		self.SWData = self.GetData()
	# SettupXL creates the Excel object, which we alter before we insert data
	def SettupXL(self): 
		self.NetBook = XL.Create()
		XL.Headers("MACHINE_NAME","A1",self.NetBook[1])
		XL.Headers("IP_ADDRESS","B1",self.NetBook[1])
		XL.Headers("ACTIVE_PORTS","D1",self.NetBook[1])
		XL.Headers("NON_ACTIVE_PORTS","E1",self.NetBook[1])
		XL.Headers("DISABLED_PORTS","F1",self.NetBook[1])
		XL.Headers("TOTAL_PORTS","G1",self.NetBook[1]) 
		
	# GetData is the method that gets solarwind data and any information we want
	def GetData(self):
		MIN = 2
		SWData = Query.Select(self.DB)
		DataList = []
		for Cap,IP in SWData.items():
			Test = Connection.TestConnectAbility(IP,self.username,self.password)
			if Test[0] == True:
				Data = Connection.Pull(Test[1],self.Command)
				ReadFile = Query.TempFileMkr(str(Cap),Data)
				ActivePorts = Query.FindOccur(ReadFile,"connected")
				InActivePorts = Query.FindOccur(ReadFile,"notconnect")
				DisabledPorts = Query.FindOccur(ReadFile,"disabled")
				TotalPorts = ActivePorts + InActivePorts + DisabledPorts
			elif Test[0] == False:
				ActivePorts = "N/A"
				InActivePorts = "N/A"
				DisabledPorts = "N/A"
				TotalPorts = "N/A"
			DBind = Query.Binder(Cap,IP,ActivePorts,InActivePorts,DisabledPorts,TotalPorts)
			XL.InsertData(self.NetBook[1],DBind,MIN)
			MIN = MIN + 1
		self.NetBook[0].save(filename = 'scplzwork.xlsx')
		
# XL class serves both as an initilizer 
# Handler role: Inserts arguments and establishes the rows
class XL:
	def Create():
		print("Creating Workbook.....")
		WB = Workbook()
		WS = WB.active
		return WB,WS
	def Headers(Header,Cell,WS):
		WS[str(Cell)] = Header
	def InsertData(WS,Data,MIN):
		WS["A" + str(MIN)] = Data['MACHINE_NAME']
		WS["B" + str(MIN)] = Data['IP_ADDRESS']
		WS["D" + str(MIN)] = Data['ACTIVE_PORTS']
		WS["E" + str(MIN)] = Data['NON_ACTIVE_PORTS']
		WS["F" + str(MIN)] = Data['DISABLED_PORTS']
		WS["G" + str(MIN)] = Data['TOTAL_PORTS']
		WS.cell(row = MIN, column = 1).alignment = Alignment(horizontal = 'left')
		WS.cell(row = MIN, column = 2).alignment = Alignment(horizontal = 'left')
		#CurCol = WS.cell(str(Col))
		#CurCol.alighnment = Alignment(horizontal = str(Direction))
	# The custom insert is becuase im too lazy to rework the insert data method above.




Main()
